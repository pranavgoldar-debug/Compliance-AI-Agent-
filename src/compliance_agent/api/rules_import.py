"""Bulk rule import — CSV / Excel.

Admin uploads a CSV or XLSX with one rule per row. Three endpoints:

  GET  /api/rules/import/template?format=csv|xlsx  - download a sample
  POST /api/rules/import/preview                   - parse + validate
  POST /api/rules/import/commit                    - persist as Staging rules

No Grok. This is the plain bulk importer for an existing compliance tracker
sheet (e.g. the Aspora Global Compliance Tracker CSV).
"""
from __future__ import annotations

import csv
import io
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from compliance_agent.api._helpers import log_activity
from compliance_agent.api.rules import _serialize_rule
from compliance_agent.api.schemas import RuleOut
from compliance_agent.auth import require_admin
from compliance_agent.db import (
    Applicability,
    Entity,
    Rule,
    RuleStatus,
    User,
    get_session,
)


router = APIRouter(prefix="/api/rules/import", tags=["rules-import"])


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------
# Canonical column names (we normalise headers to these). Anything not in
# REQUIRED + OPTIONAL is preserved on the row but ignored at commit time.
REQUIRED_COLUMNS = [
    "name",
    "jurisdiction_code",
    "category",
    "form_name",
    "authority",
    "frequency",
    "due_date_rule",
]
OPTIONAL_COLUMNS = [
    "area",
    "payment_rule",
    "applicability",
    "applicability_note",
    "source_url",
]
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

# Accept a few common header variants ('Form', 'Form name', etc.) by mapping
# anything that normalises to a known alias.
HEADER_ALIASES: dict[str, str] = {
    "rule_name": "name",
    "rule": "name",
    "jurisdiction": "jurisdiction_code",
    "country": "jurisdiction_code",
    "country_code": "jurisdiction_code",
    "form": "form_name",
    "form_no": "form_name",
    "form_number": "form_name",
    "regulator": "authority",
    "agency": "authority",
    "due_date": "due_date_rule",
    "deadline": "due_date_rule",
    "due": "due_date_rule",
    "payment": "payment_rule",
    "payment_deadline": "payment_rule",
    "applies_to": "applicability",
    "scope": "applicability",
    "source": "source_url",
    "url": "source_url",
    "regulator_url": "source_url",
    "notes": "applicability_note",
    "note": "applicability_note",
}

# jurisdiction_code aliases — accept human-friendly country names and
# 2-letter ISO codes, map them to the internal lowercase keys.
JURISDICTION_ALIASES: dict[str, str] = {
    "in": "india", "ind": "india", "india": "india",
    "gb": "uk", "uk": "uk", "united kingdom": "uk", "britain": "uk",
    "us": "us", "usa": "us", "united states": "us", "united states of america": "us",
    "eu": "eu", "european union": "eu",
    "ae": "uae", "uae": "uae", "united arab emirates": "uae",
    "sg": "singapore", "singapore": "singapore",
    "ca": "canada", "canada": "canada",
    "lt": "lithuania", "lithuania": "lithuania",
}

APPLICABILITY_ALIASES: dict[str, Applicability] = {
    "mandatory": Applicability.mandatory,
    "required": Applicability.mandatory,
    "must": Applicability.mandatory,
    "conditional": Applicability.conditional,
    "if applicable": Applicability.conditional,
    "sector-specific": Applicability.sector_specific,
    "sector specific": Applicability.sector_specific,
    "sectoral": Applicability.sector_specific,
}


def _normalise_header(raw: str) -> str:
    """Lowercase, strip, swap whitespace/dashes for underscores."""
    s = (raw or "").strip().lower()
    for ch in (" ", "-", "/", "."):
        s = s.replace(ch, "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s


def _canonical_column(header: str) -> Optional[str]:
    n = _normalise_header(header)
    if n in ALL_COLUMNS:
        return n
    return HEADER_ALIASES.get(n)


def _coerce_jurisdiction(raw: str) -> Optional[str]:
    s = (raw or "").strip().lower()
    if not s:
        return None
    return JURISDICTION_ALIASES.get(s, s if s in {
        "india", "uk", "us", "eu", "uae", "singapore", "canada", "lithuania"
    } else None)


def _coerce_applicability(raw: str) -> Optional[Applicability]:
    s = (raw or "").strip().lower()
    if not s:
        return None
    return APPLICABILITY_ALIASES.get(s)


# ---------------------------------------------------------------------------
# Parse helpers (CSV + XLSX)
# ---------------------------------------------------------------------------
def _parse_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _parse_xlsx(content: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise HTTPException(
            status_code=501,
            detail="openpyxl is not installed on the server. Upload a CSV instead.",
        ) from e
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return [], []
    header = ["" if v is None else str(v) for v in raw_rows[0]]
    data = [
        ["" if v is None else str(v) for v in row]
        for row in raw_rows[1:]
    ]
    return header, data


def _parse_upload(file: UploadFile) -> tuple[list[str], list[list[str]]]:
    name = (file.filename or "").lower()
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content)
    # Default to CSV — most tracker exports come as CSV anyway.
    return _parse_csv(content)


# ---------------------------------------------------------------------------
# Preview / commit payloads
# ---------------------------------------------------------------------------
class ParsedRow(BaseModel):
    row_number: int  # 1-based, header is row 1 so first data row is 2
    name: str = ""
    jurisdiction_code: str = ""
    category: str = ""
    area: str = ""
    form_name: str = ""
    authority: str = ""
    frequency: str = ""
    due_date_rule: str = ""
    payment_rule: Optional[str] = None
    applicability: str = "Mandatory"
    applicability_note: Optional[str] = None
    source_url: Optional[str] = None
    errors: list[str] = Field(default_factory=list)


class PreviewResponse(BaseModel):
    detected_columns: list[str]
    unknown_columns: list[str]
    missing_required: list[str]
    rows: list[ParsedRow]
    valid_count: int
    error_count: int


def _build_parsed_row(
    row_number: int,
    header: list[str],
    mapping: dict[int, str],
    values: list[str],
) -> ParsedRow:
    out: dict[str, str] = {col: "" for col in ALL_COLUMNS}
    for idx, raw in enumerate(values):
        canonical = mapping.get(idx)
        if not canonical:
            continue
        out[canonical] = (raw or "").strip()

    errors: list[str] = []

    # Required field presence.
    for col in REQUIRED_COLUMNS:
        if not out.get(col):
            errors.append(f"missing {col}")

    # Jurisdiction code normalisation.
    juris_raw = out.get("jurisdiction_code", "")
    juris = _coerce_jurisdiction(juris_raw)
    if juris_raw and not juris:
        errors.append(f"unknown jurisdiction '{juris_raw}'")
    if juris:
        out["jurisdiction_code"] = juris

    # Applicability normalisation. Empty = Mandatory.
    app_raw = out.get("applicability", "")
    app = _coerce_applicability(app_raw) if app_raw else Applicability.mandatory
    if app_raw and not app:
        errors.append(f"unknown applicability '{app_raw}' (use Mandatory / Conditional / Sector-specific)")
    out["applicability"] = (app or Applicability.mandatory).value

    return ParsedRow(
        row_number=row_number,
        name=out["name"],
        jurisdiction_code=out["jurisdiction_code"],
        category=out["category"],
        area=out["area"],
        form_name=out["form_name"],
        authority=out["authority"],
        frequency=out["frequency"],
        due_date_rule=out["due_date_rule"],
        payment_rule=out.get("payment_rule") or None,
        applicability=out["applicability"],
        applicability_note=out.get("applicability_note") or None,
        source_url=out.get("source_url") or None,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
TEMPLATE_HEADERS = [
    "name",
    "jurisdiction_code",
    "category",
    "area",
    "form_name",
    "authority",
    "frequency",
    "due_date_rule",
    "payment_rule",
    "applicability",
    "applicability_note",
    "source_url",
]

TEMPLATE_EXAMPLES: list[list[str]] = [
    [
        "India GSTR-3B monthly return",
        "india",
        "Indirect Tax",
        "GST",
        "GSTR-3B",
        "GSTN / CBIC",
        "Monthly",
        "20th of the month following the tax period",
        "Same day as filing (20th of following month)",
        "Mandatory",
        "All GST-registered entities with monthly filing frequency.",
        "https://www.gst.gov.in/",
    ],
    [
        "UAE Corporate Tax return",
        "uae",
        "Direct Tax",
        "Corporate Tax",
        "CT Return",
        "Federal Tax Authority",
        "Annual",
        "Within 9 months of the end of the relevant tax period",
        "Same day as filing",
        "Mandatory",
        "All taxable persons under UAE Corporate Tax law.",
        "https://tax.gov.ae/",
    ],
    [
        "USA Form 1095-C (ACA)",
        "us",
        "HR / Payroll",
        "Affordable Care Act",
        "Form 1095-C",
        "IRS",
        "Annual",
        "Furnish to employees by 31 March; file with IRS by 31 March (e-file)",
        "",
        "Conditional",
        "Applicable Large Employers (ALEs) with 50+ full-time employees.",
        "https://www.irs.gov/forms-pubs/about-form-1095-c",
    ],
]


@router.get("/template")
def download_template(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    _: User = Depends(require_admin),
):
    """Return a starter file with the expected headers and 3 example rows."""
    if format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as e:
            raise HTTPException(
                status_code=501,
                detail="openpyxl is not installed on the server. Use format=csv instead.",
            ) from e

        wb = Workbook()
        ws = wb.active
        ws.title = "Rules"
        ws.append(TEMPLATE_HEADERS)
        for row in TEMPLATE_EXAMPLES:
            ws.append(row)
        header_fill = PatternFill(start_color="F4F0FF", end_color="F4F0FF", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="top")
        for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 28

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": (
                    'attachment; filename="aspora-rules-import-template.xlsx"'
                )
            },
        )

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)
    for row in TEMPLATE_EXAMPLES:
        writer.writerow(row)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": (
                'attachment; filename="aspora-rules-import-template.csv"'
            )
        },
    )


@router.post("/preview", response_model=PreviewResponse)
def preview_import(
    file: UploadFile = File(...),
    _: User = Depends(require_admin),
) -> PreviewResponse:
    header, data = _parse_upload(file)
    if not header:
        raise HTTPException(
            status_code=400,
            detail="Couldn't read any header row from the file.",
        )

    # Build idx -> canonical column mapping.
    mapping: dict[int, str] = {}
    detected: list[str] = []
    unknown: list[str] = []
    for idx, raw in enumerate(header):
        canonical = _canonical_column(raw)
        if canonical:
            mapping[idx] = canonical
            detected.append(canonical)
        elif (raw or "").strip():
            unknown.append(raw)

    missing_required = [c for c in REQUIRED_COLUMNS if c not in detected]

    rows: list[ParsedRow] = []
    for offset, raw_row in enumerate(data):
        # Skip rows that are entirely blank (common in exported sheets).
        if not any((cell or "").strip() for cell in raw_row):
            continue
        rows.append(_build_parsed_row(offset + 2, header, mapping, raw_row))

    valid = sum(1 for r in rows if not r.errors)
    return PreviewResponse(
        detected_columns=detected,
        unknown_columns=unknown,
        missing_required=missing_required,
        rows=rows,
        valid_count=valid,
        error_count=len(rows) - valid,
    )


class CommitRow(BaseModel):
    name: str
    jurisdiction_code: str
    category: str
    area: str = ""
    form_name: str
    authority: str
    frequency: str
    due_date_rule: str
    payment_rule: Optional[str] = None
    applicability: Applicability = Applicability.mandatory
    applicability_note: Optional[str] = None
    source_url: Optional[str] = None


class CommitPayload(BaseModel):
    rows: list[CommitRow]
    entity_ids: list[int] = []
    status: RuleStatus = RuleStatus.staging


class CommitResponse(BaseModel):
    created: list[RuleOut]


@router.post("/commit", response_model=CommitResponse)
def commit_import(
    payload: CommitPayload,
    db: Session = Depends(get_session),
    user: User = Depends(require_admin),
) -> CommitResponse:
    if not payload.rows:
        raise HTTPException(status_code=400, detail="Provide at least one row to import.")

    entities: list[Entity] = []
    if payload.entity_ids:
        entities = (
            db.execute(select(Entity).where(Entity.id.in_(payload.entity_ids)))
            .scalars()
            .all()
        )

    created: list[Rule] = []
    for r in payload.rows:
        rule = Rule(
            name=r.name,
            jurisdiction_code=r.jurisdiction_code,
            category=r.category,
            area=r.area,
            form_name=r.form_name,
            authority=r.authority,
            frequency=r.frequency,
            due_date_rule=r.due_date_rule,
            payment_rule=r.payment_rule,
            applicability=r.applicability,
            applicability_note=r.applicability_note,
            source_url=r.source_url,
            status=payload.status,
            created_by_id=user.id,
        )
        # Attach only entities whose jurisdiction matches the rule.
        # Cross-jurisdiction attachments are almost always a mistake, and the
        # admin can still attach them after the fact from the rules table.
        rule.entities = [e for e in entities if e.jurisdiction_code == r.jurisdiction_code]
        db.add(rule)
        created.append(rule)
    db.flush()

    log_activity(
        db,
        actor_id=user.id,
        action="rule.bulk_imported",
        target_type="rule",
        payload={
            "count": len(created),
            "entity_ids": payload.entity_ids,
            "status": payload.status.value,
            "source": "csv_import",
        },
    )
    db.commit()
    for rule in created:
        db.refresh(rule)
    return CommitResponse(created=[_serialize_rule(r) for r in created])
