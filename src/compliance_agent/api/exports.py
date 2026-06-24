"""Unified CSV / Excel export endpoints for the major list views.

Each list-view export accepts the same filter query params as its read API
and returns either text/csv or application/vnd.openxmlformats… depending on
the `format` query param. openpyxl is imported lazily so a missing dep
falls back to a 501.

  GET /api/exports/obligations?format=csv|xlsx&entity_id=…&status=…
  GET /api/exports/entities?format=csv|xlsx
  GET /api/exports/rules?format=csv|xlsx&status=…
  GET /api/exports/documents?format=csv|xlsx&entity_id=…
"""
from __future__ import annotations

import csv
import io
from datetime import date
from typing import Iterable, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from compliance_agent.api._helpers import days_remaining, is_overdue, is_in_alert_window
from compliance_agent.auth import get_current_user
from compliance_agent.classification import keep_function
from compliance_agent.db import (
    Document,
    EffortBand,
    Entity,
    Obligation,
    ObligationStatus,
    Rule,
    RuleStatus,
    User,
    get_session,
)


router = APIRouter(prefix="/api/exports", tags=["exports"])


# ---------------------------------------------------------------------------
# Format helpers
# ---------------------------------------------------------------------------
def _stream_csv(filename: str, headers: list[str], rows: Iterable[list[object]]) -> StreamingResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(["" if v is None else v for v in r])
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


def _stream_xlsx(
    filename: str,
    sheet_name: str,
    headers: list[str],
    rows: Iterable[list[object]],
) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise HTTPException(
            status_code=501,
            detail="openpyxl is not installed on the server. Use format=csv instead.",
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Export"

    header_fill = PatternFill(start_color="F4F0FF", end_color="F4F0FF", fill_type="solid")
    header_font = Font(bold=True, color="3D1A7A")

    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    rows = list(rows)
    for r in rows:
        ws.append(["" if v is None else v for v in r])

    # Autosize columns based on data length (capped to keep things sane).
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in rows:
            v = r[col_idx - 1]
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def _emit(
    format: str,
    filename: str,
    sheet_name: str,
    headers: list[str],
    rows: Iterable[list[object]],
) -> StreamingResponse:
    fmt = (format or "csv").lower()
    if fmt == "csv":
        return _stream_csv(filename, headers, rows)
    if fmt in ("xlsx", "excel"):
        return _stream_xlsx(filename, sheet_name, headers, rows)
    raise HTTPException(status_code=400, detail="Unknown format. Use 'csv' or 'xlsx'.")


# ---------------------------------------------------------------------------
# Obligations export — same filters as /api/obligations
# ---------------------------------------------------------------------------
@router.get("/obligations")
def export_obligations(
    format: str = Query("csv"),
    entity_id: Optional[int] = Query(None),
    status: Optional[ObligationStatus] = Query(None),
    due_from: Optional[date] = Query(None),
    due_to: Optional[date] = Query(None),
    jurisdiction_code: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    db: Session = Depends(get_session),
    _: User = Depends(get_current_user),
):
    stmt = (
        select(Obligation)
        .options(
            joinedload(Obligation.rule),
            joinedload(Obligation.entity),
            joinedload(Obligation.assignee),
        )
        .order_by(Obligation.due_date.asc())
    )
    if entity_id is not None:
        stmt = stmt.where(Obligation.entity_id == entity_id)
    if status is not None:
        stmt = stmt.where(Obligation.status == status)
    if due_from is not None:
        stmt = stmt.where(Obligation.due_date >= due_from)
    if due_to is not None:
        stmt = stmt.where(Obligation.due_date <= due_to)

    items = db.execute(stmt).scalars().unique().all()
    if jurisdiction_code:
        items = [o for o in items if o.rule.jurisdiction_code == jurisdiction_code]
    if category:
        items = [o for o in items if o.rule.category == category]
    # FINANCE_ONLY switch: keep exports consistent with the in-app lists.
    items = [
        o
        for o in items
        if o.rule is None
        or keep_function(o.rule.category, o.rule.area, o.rule.responsible_function)
    ]

    headers = [
        "Due date",
        "Entity",
        "Jurisdiction",
        "Form",
        "Authority",
        "Category",
        "Frequency",
        "Period",
        "Effort band",
        "Status",
        "Assignee",
        "Days remaining",
        "Filing reference",
        "Payment amount",
        "Payment reference",
        "Notes",
    ]
    rows = []
    for o in items:
        band = o.effort_band or EffortBand.w4
        rows.append(
            [
                o.due_date.isoformat(),
                o.entity.name if o.entity else "",
                o.entity.jurisdiction_code if o.entity else "",
                o.rule.form_name if o.rule else "",
                o.rule.authority if o.rule else "",
                o.rule.category if o.rule else "",
                o.rule.frequency if o.rule else "",
                o.period_label or "",
                band.value if band else "",
                "Overdue" if is_overdue(o.due_date, o.status) else o.status.value.replace("_", " "),
                (o.assignee.full_name or o.assignee.email) if o.assignee else "",
                days_remaining(o.due_date),
                o.filing_reference or "",
                o.payment_amount or "",
                o.payment_reference or "",
                (o.notes or "").replace("\n", " "),
            ]
        )

    return _emit(
        format,
        filename="aspora-obligations",
        sheet_name="Obligations",
        headers=headers,
        rows=rows,
    )


# ---------------------------------------------------------------------------
# Entities export
# ---------------------------------------------------------------------------
@router.get("/entities")
def export_entities(
    format: str = Query("csv"),
    jurisdiction_code: Optional[str] = Query(None),
    db: Session = Depends(get_session),
    _: User = Depends(get_current_user),
):
    stmt = (
        select(Entity)
        .order_by(Entity.name)
    )
    if jurisdiction_code:
        stmt = stmt.where(Entity.jurisdiction_code == jurisdiction_code)

    entities = db.execute(stmt).scalars().unique().all()

    # Precompute counts in one go so we don't N+1.
    from sqlalchemy import func
    from compliance_agent.api._helpers import today as _today_fn
    today_d = _today_fn()
    rows = []
    for e in entities:
        active = db.execute(
            select(func.count(Obligation.id)).where(
                Obligation.entity_id == e.id,
                Obligation.status.notin_(
                    [ObligationStatus.completed, ObligationStatus.not_applicable]
                ),
            )
        ).scalar_one()
        overdue = db.execute(
            select(func.count(Obligation.id)).where(
                Obligation.entity_id == e.id,
                Obligation.due_date < today_d,
                Obligation.status.notin_(
                    [ObligationStatus.completed, ObligationStatus.not_applicable]
                ),
            )
        ).scalar_one()
        rows.append(
            [
                e.name,
                e.legal_type,
                e.jurisdiction_code,
                e.registration_number or "",
                e.incorporation_date.isoformat() if e.incorporation_date else "",
                e.fiscal_year_end or "",
                active,
                overdue,
            ]
        )

    headers = [
        "Name",
        "Type",
        "Jurisdiction",
        "Registration #",
        "Incorporation date",
        "Fiscal year end",
        "Active obligations",
        "Overdue",
    ]
    return _emit(format, "aspora-entities", "Entities", headers, rows)


# ---------------------------------------------------------------------------
# Rules export
# ---------------------------------------------------------------------------
@router.get("/rules")
def export_rules(
    format: str = Query("csv"),
    status: Optional[RuleStatus] = Query(None),
    jurisdiction_code: Optional[str] = Query(None),
    in_review: Optional[bool] = Query(
        None,
        description="When true, only rules sent to Review & Assign "
        "(sent_to_review is True) — matches the For Action tab, so the export "
        "excludes freshly-discovered drafts.",
    ),
    db: Session = Depends(get_session),
    _: User = Depends(get_current_user),
):
    stmt = select(Rule).order_by(Rule.jurisdiction_code, Rule.form_name)
    if status is not None:
        stmt = stmt.where(Rule.status == status)
    if jurisdiction_code:
        stmt = stmt.where(Rule.jurisdiction_code == jurisdiction_code)
    if in_review:
        stmt = stmt.where(Rule.sent_to_review.is_(True))
    rules = db.execute(stmt).scalars().all()
    # FINANCE_ONLY switch: keep the rules export Finance-only too.
    rules = [r for r in rules if keep_function(r.category, r.area, r.responsible_function)]

    headers = [
        "Jurisdiction",
        "Form",
        "Authority",
        "Category",
        "Area",
        "Frequency",
        "Due-date rule",
        "Payment rule",
        "Applicability",
        "Status",
        "Updated at",
    ]
    rows = [
        [
            r.jurisdiction_code,
            r.form_name,
            r.authority,
            r.category,
            r.area,
            r.frequency,
            (r.due_date_rule or "").replace("\n", " "),
            (r.payment_rule or "").replace("\n", " "),
            r.applicability.value if r.applicability else "",
            r.status.value if r.status else "",
            r.updated_at.isoformat() if r.updated_at else "",
        ]
        for r in rules
    ]
    return _emit(format, "aspora-rules", "Rules", headers, rows)


# ---------------------------------------------------------------------------
# Documents export
# ---------------------------------------------------------------------------
@router.get("/documents")
def export_documents(
    format: str = Query("csv"),
    entity_id: Optional[int] = Query(None),
    db: Session = Depends(get_session),
    _: User = Depends(get_current_user),
):
    stmt = (
        select(Document)
        .options(
            joinedload(Document.entity),
            joinedload(Document.uploaded_by),
            joinedload(Document.obligation).joinedload(Obligation.rule),
        )
        .order_by(Document.created_at.desc())
    )
    if entity_id is not None:
        stmt = stmt.where(Document.entity_id == entity_id)
    docs = db.execute(stmt).scalars().unique().all()

    headers = [
        "Filename",
        "Entity",
        "Linked obligation",
        "Category",
        "Tags",
        "Size (KB)",
        "Content type",
        "Uploaded by",
        "Uploaded at",
    ]
    rows = [
        [
            d.filename,
            d.entity.name if d.entity else "",
            (d.obligation.rule.form_name if d.obligation and d.obligation.rule else ""),
            d.category.value,
            d.tags or "",
            round(d.size_bytes / 1024, 1),
            d.content_type or "",
            (d.uploaded_by.full_name if d.uploaded_by else "") or "",
            d.created_at.isoformat(),
        ]
        for d in docs
    ]
    return _emit(format, "aspora-documents", "Documents", headers, rows)


# Re-export the suppress unused import.
_ = is_in_alert_window
